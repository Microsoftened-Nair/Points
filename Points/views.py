from django.shortcuts import render
from django.http import HttpResponse
import openpyxl as op

def addm(request):
    err_msg=""
    wb = op.load_workbook(filename="Points/Points.xlsx")
    ws = wb.active
    megh = ws['A1']
    rishi = ws['B1']
    mp = megh.value
    rp = rishi.value
    if rp > 0:
        ws['B1'] = rp - 1
        ws['A1'] = mp + 1
        megh = ws['A1']
        rishi = ws['B1']
        mp = megh.value
        rp = rishi.value
    else:
        err_msg = 'Insufficient balance'
    wb.save(filename="Points/Points.xlsx")

    return HttpResponse("""<html><script>window.location.replace('/');</script></html>""")

def addr(request):
    err_msg=""
    wb = op.load_workbook(filename="Points/Points.xlsx")
    ws = wb.active
    megh = ws['A1']
    rishi = ws['B1']
    mp = megh.value
    rp = rishi.value

    if mp > 0:
        ws['A1'] = mp - 1
        ws['B1'] = rp + 1
        megh = ws['A1']
        rishi = ws['B1']
        mp = megh.value
        rp = rishi.value
    else:
        err_msg = 'Insufficient balance'

    wb.save(filename='Points/Points.xlsx')

    return HttpResponse("""<html><script>window.location.replace('/');</script></html>""")


def index(request):
    err_msg=""
    wb = op.load_workbook(filename="Points/Points.xlsx")
    ws = wb.active
    megh = ws['A1']
    rishi = ws['B1']
    mp = megh.value
    rp = rishi.value
    return render(request, 'index.html',{
    'rp':rp,
    'mp':mp,
    'err_msg':err_msg
    })